import openpyxl
from io import BytesIO
from django.http import HttpResponse
from django.db import transaction
from .forms import MeteoDataUploadForm
from django.views.generic import TemplateView, ListView, DetailView, FormView, CreateView, View
from django.urls import reverse, reverse_lazy
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Count, Avg, Min, Max, Q
from django.core.paginator import Paginator
from .forms import ExpeditionForm, StationForm, CTDProfileForm
from .models import (
    Expedition, Station, Sample, MeteoData, CarbonData, 
    IonicCompositionData, PigmentsData, OxymetrData, 
    NutrientsData, PHMeasurement, Probe, CTDData, CTDProfile
)


class ComingSoonView(TemplateView):
    template_name = 'oceanography/coming_soon.html'

class HomeView(TemplateView):
    template_name = 'oceanography/home.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Основная статистика
        context['expeditions_count'] = Expedition.objects.count()
        context['stations_count'] = Station.objects.count()
        context['samples_count'] = Sample.objects.count()
        context['ctd_data_count'] = CTDData.objects.count()
        
        # Последние экспедиции
        context['recent_expeditions'] = Expedition.objects.select_related().order_by('-start_date')[:5]
        
        # Последние станции с данными об экспедициях
        context['recent_stations'] = Station.objects.select_related('expedition').order_by('-datetime')[:10]
        
        # Последние пробы
        context['recent_samples'] = Sample.objects.select_related(
            'station', 'station__expedition'
        ).order_by('-datetime')[:10]
        
        # Статистика по типам данных
        context['data_stats'] = {
            'meteo': MeteoData.objects.count(),
            'carbon': CarbonData.objects.count(),
            'ionic': IonicCompositionData.objects.count(),
            'pigments': PigmentsData.objects.count(),
            'nutrients': NutrientsData.objects.count(),
            'ph': PHMeasurement.objects.count(),
        }
        
        # Последняя активность
        latest_station = Station.objects.order_by('-datetime').first()
        if latest_station:
            context['latest_activity'] = {
                'type': 'станция',
                'name': latest_station.station_name,
                'expedition': latest_station.expedition.platform,
                'date': latest_station.datetime
            }
        else:
            latest_expedition = Expedition.objects.order_by('-start_date').first()
            if latest_expedition:
                context['latest_activity'] = {
                    'type': 'экспедиция',
                    'name': latest_expedition.platform,
                    'date': latest_expedition.start_date
                }
        
        return context

class ExpeditionListView(ListView):
    model = Expedition
    template_name = 'oceanography/expedition_list.html'
    context_object_name = 'expeditions'
    paginate_by = 20
    
    def get_queryset(self):
        return Expedition.objects.all().prefetch_related('stations')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': '', 'name': 'Экспедиции'}
        ]
        return context

class ExpeditionCreateView(CreateView):
    model = Expedition
    form_class = ExpeditionForm
    template_name = 'oceanography/expedition_form.html'
    success_url = reverse_lazy('oceanography:expedition_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Экспедиция "{self.object.platform}" успешно создана!')
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:expedition_list'), 'name': 'Экспедиции'},
            {'url': '', 'name': 'Создание экспедиции'}
        ]
        return context

class ExpeditionDetailView(DetailView):
    model = Expedition
    template_name = 'oceanography/expedition_detail.html'
    context_object_name = 'expedition'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        expedition = self.get_object()
        
        # Оптимизируем запросы к базе данных
        context['stations'] = expedition.stations.all().prefetch_related('samples')
        context['samples_count'] = Sample.objects.filter(station__expedition=expedition).count()
        
        # Расчет длительности в днях
        if expedition.start_date and expedition.end_date:
            delta = expedition.end_date - expedition.start_date
            context['duration_days'] = delta.days
        else:
            context['duration_days'] = None
        
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:expedition_list'), 'name': 'Экспедиции'},
            {'url': '', 'name': f'{expedition.platform} ({expedition.start_date.year})'}
        ]
        return context

class StationSingleCreateView(CreateView):
    """Форма для добавления одной станции с автоматическим созданием двух проб"""
    model = Station
    form_class = StationForm
    template_name = 'oceanography/station_single_create.html'
    
    def form_valid(self, form):
        expedition_id = self.kwargs.get('expedition_id')
        form.instance.expedition_id = expedition_id
        
        # Проверка уникальности даты/времени станции в экспедиции
        if Station.objects.filter(
            expedition_id=expedition_id, 
            datetime=form.instance.datetime
        ).exists():
            form.add_error(
                'datetime', 
                f'Станция с датой/времением "{form.instance.datetime}" уже существует в этой экспедиции'
            )
            return self.form_invalid(form)
        
        # Сохраняем станцию
        response = super().form_valid(form)
        
        # Автоматически создаем две пробы
        Sample.objects.create(
            station=self.object,
            datetime=self.object.datetime,
            sampling_depth='0.0',
            comment='Поверхностная проба'
        )
        Sample.objects.create(
            station=self.object,
            datetime=self.object.datetime,
            sampling_depth='дно',
            comment='Придонная проба'
        )
        
        messages.success(
            self.request, 
            f'Станция "{form.instance.station_name}" успешно создана с двумя автоматическими пробами (поверхность и дно)!'
        )
        
        return response
    
    def get_success_url(self):
        expedition_id = self.kwargs.get('expedition_id')
        return reverse('oceanography:expedition_detail', kwargs={'pk': expedition_id})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        expedition_id = self.kwargs.get('expedition_id')
        
        try:
            context['expedition'] = Expedition.objects.get(pk=expedition_id)
        except Expedition.DoesNotExist:
            context['expedition'] = None
        
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:expedition_list'), 'name': 'Экспедиции'},
            {'url': reverse('oceanography:expedition_detail', kwargs={'pk': expedition_id}), 
             'name': f'Экспедиция {expedition_id}'},
            {'url': '', 'name': 'Добавление станции'}
        ]
        return context



# ============================================================================
# ПРЕДСТАВЛЕНИЯ ДЛЯ ПРОСМОТРА ВСЕХ ДАННЫХ
# ============================================================================

class DataOverviewView(TemplateView):
    """Обзор всех данных в системе"""
    template_name = 'oceanography/data_overview.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Статистика по всем основным таблицам
        context['stats'] = {
            'expeditions': {
                'count': Expedition.objects.count(),
                'recent': Expedition.objects.order_by('-start_date')[:5],
                'total_stations': Station.objects.count(),
            },
            'stations': {
                'count': Station.objects.count(),
                'with_samples': Station.objects.filter(samples__isnull=False).distinct().count(),
                'recent': Station.objects.select_related('expedition').order_by('-datetime')[:10],
            },
            'samples': {
                'count': Sample.objects.count(),
                'recent': Sample.objects.select_related('station', 'station__expedition').order_by('-datetime')[:10],
            },
            'meteo_data': {
                'count': MeteoData.objects.count(),
                'recent': MeteoData.objects.select_related('sample', 'sample__station', 'sample__station__expedition')[:10],
            },
            'carbon_data': {
                'count': CarbonData.objects.count(),
                'recent': CarbonData.objects.select_related('sample', 'sample__station', 'sample__station__expedition')[:10],
            },
            'ionic_data': {
                'count': IonicCompositionData.objects.count(),
                'recent': IonicCompositionData.objects.select_related('sample', 'sample__station', 'sample__station__expedition')[:10],
            },
            'pigments_data': {
                'count': PigmentsData.objects.count(),
                'recent': PigmentsData.objects.select_related('sample', 'sample__station', 'sample__station__expedition')[:10],
            },
            'oxymetr_data': {
                'count': OxymetrData.objects.count(),
                'recent': OxymetrData.objects.select_related('sample', 'sample__station', 'sample__station__expedition')[:10],
            },
            'nutrients_data': {
                'count': NutrientsData.objects.count(),
                'recent': NutrientsData.objects.select_related('sample', 'sample__station', 'sample__station__expedition')[:10],
            },
            'ph_measurements': {
                'count': PHMeasurement.objects.count(),
                'recent': PHMeasurement.objects.select_related('sample', 'sample__station', 'sample__station__expedition')[:10],
            },
            'probes': {
                'count': Probe.objects.count(),
                'recent': Probe.objects.all()[:10],
            },
            'ctd_data': {
                'count': CTDData.objects.count(),
                'stations_with_ctd': CTDData.objects.values('sample__station').distinct().count(),
                'recent': CTDData.objects.select_related('sample', 'sample__station', 'sample__station__expedition', 'probe')[:10],
            },
        }
        
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': '', 'name': 'Обзор данных'}
        ]
        
        return context

class ExpeditionDataView(ListView):
    """Детальный просмотр всех экспедиций"""
    model = Expedition
    template_name = 'oceanography/data_expeditions.html'
    context_object_name = 'expeditions'
    paginate_by = 20
    
    def get_queryset(self):
        return Expedition.objects.annotate(
            stations_count=Count('stations'),
            samples_count=Count('stations__samples')
        ).order_by('-start_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:data_overview'), 'name': 'Обзор данных'},
            {'url': '', 'name': 'Все экспедиции'}
        ]
        return context

class StationDataView(ListView):
    """Детальный просмотр всех станций"""
    model = Station
    template_name = 'oceanography/data_stations.html'
    context_object_name = 'stations'
    paginate_by = 50
    
    def get_queryset(self):
        return Station.objects.select_related('expedition').annotate(
            samples_count=Count('samples')
        ).order_by('-datetime')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:data_overview'), 'name': 'Обзор данных'},
            {'url': '', 'name': 'Все станции'}
        ]
        return context

class SampleDataView(ListView):
    """Детальный просмотр всех проб"""
    model = Sample
    template_name = 'oceanography/data_samples.html'
    context_object_name = 'samples'
    paginate_by = 50
    
    def get_queryset(self):
        return Sample.objects.select_related(
            'station', 'station__expedition'
        ).prefetch_related(
            'meteo_data', 'carbon_data', 'ionic_data', 
            'pigments_data', 'nutrients_data', 'ph_measurements'
        ).order_by('-datetime')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:data_overview'), 'name': 'Обзор данных'},
            {'url': '', 'name': 'Все пробы'}
        ]
        return context

class MeteoDataView(ListView):
    """Детальный просмотр всех метеоданных"""
    model = MeteoData
    template_name = 'oceanography/data_meteo.html'
    context_object_name = 'meteo_data'
    paginate_by = 50
    
    def get_queryset(self):
        return MeteoData.objects.select_related(
            'sample', 'sample__station', 'sample__station__expedition'
        ).order_by('-sample__datetime')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:data_overview'), 'name': 'Обзор данных'},
            {'url': '', 'name': 'Метеоданные'}
        ]
        return context

class CarbonDataView(ListView):
    """Детальный просмотр всех данных по углероду"""
    model = CarbonData
    template_name = 'oceanography/data_carbon.html'
    context_object_name = 'carbon_data'
    paginate_by = 50
    
    def get_queryset(self):
        return CarbonData.objects.select_related(
            'sample', 'sample__station', 'sample__station__expedition'
        ).order_by('-sample__datetime')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:data_overview'), 'name': 'Обзор данных'},
            {'url': '', 'name': 'Данные по углероду'}
        ]
        return context

class IonicDataView(ListView):
    """Детальный просмотр всех данных по ионному составу"""
    model = IonicCompositionData
    template_name = 'oceanography/data_ionic.html'
    context_object_name = 'ionic_data'
    paginate_by = 50
    
    def get_queryset(self):
        return IonicCompositionData.objects.select_related(
            'sample', 'sample__station', 'sample__station__expedition'
        ).order_by('-sample__datetime')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:data_overview'), 'name': 'Обзор данных'},
            {'url': '', 'name': 'Данные по ионному составу'}
        ]
        return context

class PigmentsDataView(ListView):
    """Детальный просмотр всех данных по пигментам"""
    model = PigmentsData
    template_name = 'oceanography/data_pigments.html'
    context_object_name = 'pigments_data'
    paginate_by = 50
    
    def get_queryset(self):
        return PigmentsData.objects.select_related(
            'sample', 'sample__station', 'sample__station__expedition'
        ).order_by('-sample__datetime')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:data_overview'), 'name': 'Обзор данных'},
            {'url': '', 'name': 'Данные по пигментам'}
        ]
        return context

class OxymetrDataView(ListView):
    """Детальный просмотр всех данных оксиметра"""
    model = OxymetrData
    template_name = 'oceanography/data_oxymetr.html'
    context_object_name = 'oxymetr_data'
    paginate_by = 50
    
    def get_queryset(self):
        return OxymetrData.objects.select_related(
            'sample', 'sample__station', 'sample__station__expedition'
        ).order_by('-sample__datetime')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:data_overview'), 'name': 'Обзор данных'},
            {'url': '', 'name': 'Данные оксиметра'}
        ]
        return context

class NutrientsDataView(ListView):
    """Детальный просмотр всех данных по биогенным элементам"""
    model = NutrientsData
    template_name = 'oceanography/data_nutrients.html'
    context_object_name = 'nutrients_data'
    paginate_by = 50
    
    def get_queryset(self):
        return NutrientsData.objects.select_related(
            'sample', 'sample__station', 'sample__station__expedition'
        ).order_by('-sample__datetime')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:data_overview'), 'name': 'Обзор данных'},
            {'url': '', 'name': 'Данные по биогенным элементам'}
        ]
        return context

class PHDataView(ListView):
    """Детальный просмотр всех измерений pH"""
    model = PHMeasurement
    template_name = 'oceanography/data_ph.html'
    context_object_name = 'ph_measurements'
    paginate_by = 50
    
    def get_queryset(self):
        return PHMeasurement.objects.select_related(
            'sample', 'sample__station', 'sample__station__expedition'
        ).order_by('-sample__datetime')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:data_overview'), 'name': 'Обзор данных'},
            {'url': '', 'name': 'Измерения pH'}
        ]
        return context

class ProbeDataView(ListView):
    """Детальный просмотр всех зондов"""
    model = Probe
    template_name = 'oceanography/data_probes.html'
    context_object_name = 'probes'
    paginate_by = 50
    
    def get_queryset(self):
        return Probe.objects.annotate(
            ctd_measurements_count=Count('ctd_measurements')
        ).order_by('probe_name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:data_overview'), 'name': 'Обзор данных'},
            {'url': '', 'name': 'Зонды'}
        ]
        return context

class CTDDataView(ListView):
    """Детальный просмотр всех CTD данных"""
    model = CTDData
    template_name = 'oceanography/data_ctd.html'
    context_object_name = 'ctd_data'
    paginate_by = 50
    
    def get_queryset(self):
        return CTDData.objects.select_related(
            'sample', 'sample__station', 'sample__station__expedition', 'probe'
        ).order_by('-sample__datetime')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:data_overview'), 'name': 'Обзор данных'},
            {'url': '', 'name': 'CTD данные'}
        ]
        return context

# Добавить в views.py

class CTDProfileListView(ListView):
    """Список всех CTD профилей"""
    model = CTDProfile
    template_name = 'oceanography/ctd_profile_list.html'
    context_object_name = 'profiles'
    paginate_by = 20
    
    def get_queryset(self):
        return CTDProfile.objects.select_related(
            'station', 'station__expedition', 'probe'
        ).prefetch_related('measurements').order_by('-start_datetime')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': '', 'name': 'CTD профили'}
        ]
        return context

class CTDProfileDetailView(DetailView):
    """Детальный просмотр CTD профиля"""
    model = CTDProfile
    template_name = 'oceanography/ctd_profile_detail.html'
    context_object_name = 'profile'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        profile = self.get_object()
        
        # Получаем измерения профиля
        context['measurements'] = profile.measurements.all().order_by('depth_m')
        context['measurements_count'] = context['measurements'].count()
        
        # Статистика по измерениям
        if context['measurements']:
            context['depth_range'] = {
                'min': context['measurements'].aggregate(Min('depth_m'))['depth_m__min'],
                'max': context['measurements'].aggregate(Max('depth_m'))['depth_m__max']
            }
            context['temp_range'] = {
                'min': context['measurements'].aggregate(Min('temp_c'))['temp_c__min'],
                'max': context['measurements'].aggregate(Max('temp_c'))['temp_c__max']
            }
            context['salinity_range'] = {
                'min': context['measurements'].aggregate(Min('salinity_psu'))['salinity_psu__min'],
                'max': context['measurements'].aggregate(Max('salinity_psu'))['salinity_psu__max']
            }
        
        context['breadcrumbs'] = [
            {'url': reverse('oceanography:home'), 'name': 'Главная'},
            {'url': reverse('oceanography:ctd_profile_list'), 'name': 'CTD профили'},
            {'url': '', 'name': f'Профиль {profile.profile_id}'}
        ]
        return context

class CTDProfileCreateView(CreateView):
    """Создание нового CTD профиля"""
    model = CTDProfile
    form_class = CTDProfileForm
    template_name = 'oceanography/ctd_profile_form.html'
    
    def get_initial(self):
        """Установка начальных значений, если передан station_id"""
        initial = super().get_initial()
        station_id = self.kwargs.get('station_id')
        if station_id:
            initial['station'] = get_object_or_404(Station, pk=station_id)
        return initial
    
    def form_valid(self, form):
        # Базовая валидация - проверка что start_datetime раньше end_datetime
        if form.cleaned_data['start_datetime'] > form.cleaned_data['end_datetime']:
            form.add_error('end_datetime', 'Время окончания должно быть позже времени начала')
            return self.form_invalid(form)
        
        response = super().form_valid(form)
        messages.success(self.request, f'CTD профиль успешно создан! Файл данных можно будет обработать позже.')
        return response
    
    def get_success_url(self):
        return reverse('oceanography:ctd_profile_detail', kwargs={'pk': self.object.pk})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        station_id = self.kwargs.get('station_id')
        
        if station_id:
            context['station'] = get_object_or_404(Station, pk=station_id)
            context['breadcrumbs'] = [
                {'url': reverse('oceanography:home'), 'name': 'Главная'},
                {'url': reverse('oceanography:expedition_list'), 'name': 'Экспедиции'},
                {'url': reverse('oceanography:expedition_detail', 
                              kwargs={'pk': context['station'].expedition.pk}), 
                 'name': f'Экспедиция {context["station"].expedition.pk}'},
                {'url': '', 'name': 'Добавление CTD профиля'}
            ]
        else:
            context['breadcrumbs'] = [
                {'url': reverse('oceanography:home'), 'name': 'Главная'},
                {'url': reverse('oceanography:ctd_profile_list'), 'name': 'CTD профили'},
                {'url': '', 'name': 'Создание профиля'}
            ]
        
        return context

class MeteoExcelUploadView(View):
    """Массовое добавление метеоданных через Excel"""
    template_name = 'oceanography/meteo_excel_upload.html'
    
    def get(self, request, *args, **kwargs):
        expedition_id = self.kwargs.get('expedition_id')
        expedition = get_object_or_404(Expedition, pk=expedition_id)
        
        # Получаем пробы без метеоданных
        samples_without_meteo = Sample.objects.filter(
            station__expedition=expedition
        ).exclude(
            meteo_data__isnull=False
        ).select_related('station').order_by('station__station_name', 'datetime')
        
        context = {
            'expedition': expedition,
            'samples_without_meteo': samples_without_meteo,
            'samples_count': samples_without_meteo.count(),
            'breadcrumbs': [
                {'url': reverse('oceanography:home'), 'name': 'Главная'},
                {'url': reverse('oceanography:expedition_list'), 'name': 'Экспедиции'},
                {'url': reverse('oceanography:expedition_detail', kwargs={'pk': expedition_id}), 
                 'name': f'Экспедиция {expedition.platform}'},
                {'url': '', 'name': 'Массовое добавление метеоданных (Excel)'}
            ]
        }
        return render(request, self.template_name, context)
    
    def post(self, request, *args, **kwargs):
        expedition_id = self.kwargs.get('expedition_id')
        expedition = get_object_or_404(Expedition, pk=expedition_id)
        
        action = request.POST.get('action')
        
        if action == 'download_template':
            return self.download_template(expedition)
        elif action == 'upload_data':
            return self.upload_data(request, expedition)
        
        return redirect('oceanography:add_meteo_excel', expedition_id=self.kwargs.get('expedition_id'))
    
    def download_template(self, expedition):
        """Генерация шаблона Excel для массового добавления метеоданных"""
        # Получаем пробы без метеоданных
        samples = Sample.objects.filter(
            station__expedition=expedition
        ).exclude(
            meteo_data__isnull=False
        ).select_related('station').order_by('station__station_name', 'datetime')
        
        # Создаем новую книгу Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Метеоданные"
        
        # Заголовки (машинные и человекочитаемые)
        headers_machine = [
            'sample_id', 'station_name', 'sample_datetime', 'sampling_depth',
            't_air_c', 'humidity_percent', 'wind_speed_m_s', 
            'wind_direction', 'pressure_hpa'
        ]
        
        headers_human = [
            'ID пробы*', 'Название станции*', 'Дата и время пробы*', 'Горизонт отбора*',
            'Температура воздуха (°C)', 'Влажность (%)', 'Скорость ветра (м/с)', 
            'Направление ветра', 'Атмосферное давление (гПа)'
        ]
        
        # Записываем две строки заголовков
        for col, header in enumerate(headers_machine, 1):
            ws.cell(row=1, column=col, value=header)
        for col, header in enumerate(headers_human, 1):
            ws.cell(row=2, column=col, value=header)
        
        # Заполняем данные о пробах
        for row, sample in enumerate(samples, 3):
            ws.cell(row=row, column=1, value=sample.sample_id)
            ws.cell(row=row, column=2, value=sample.station.station_name)
            ws.cell(row=row, column=3, value=sample.datetime.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=4, value=sample.sampling_depth)
            # Остальные колонки (5-9) оставляем пустыми для заполнения пользователем
        
        # Стилизация заголовков
        for row in [1, 2]:
            for cell in ws[row]:
                cell.font = openpyxl.styles.Font(bold=True)
                if row == 1:
                    # Машинные заголовки - серый фон
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
                    )
                    cell.font = openpyxl.styles.Font(bold=True, color="666666")
                else:
                    # Человекочитаемые заголовки - синий фон
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color="E7F1FF", end_color="E7F1FF", fill_type="solid"
                    )
        
        # Добавляем пояснения на отдельный лист
        ws_help = wb.create_sheet("Инструкция")
        ws_help.append(["Поле", "Описание", "Обязательное", "Пример"])
        instructions = [
            ["sample_id", "Уникальный ID пробы (не изменять!)", "Да", "123"],
            ["station_name", "Название станции (информационное)", "Да", "СТАНЦИЯ_1"],
            ["sample_datetime", "Дата и время пробы", "Да", "2024-07-15 10:00:00"],
            ["sampling_depth", "Горизонт отбора пробы", "Да", "0.0, дно"],
            ["t_air_c", "Температура воздуха в °C", "Нет", "25.5"],
            ["humidity_percent", "Относительная влажность в %", "Нет", "65.0"],
            ["wind_speed_m_s", "Скорость ветра в м/с", "Нет", "3.5"],
            ["wind_direction", "Направление ветра в градусах", "Нет", "180"],
            ["pressure_hpa", "Атмосферное давление в гПа", "Нет", "1013.2"],
        ]
        
        for instruction in instructions:
            ws_help.append(instruction)
        
        # Автоподбор ширины колонок для основного листа
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Автоподбор для листа инструкции
        for column in ws_help.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws_help.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Создаем HTTP response
        filename = f"meteo_template_{expedition.platform}_{expedition.start_date.year}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    def upload_data(self, request, expedition):
        """Обработка загруженного Excel файла с метеоданными"""
        if 'excel_file' not in request.FILES:
            messages.error(request, 'Пожалуйста, выберите файл для загрузки')
            return redirect('oceanography:add_meteo_excel', expedition_id=expedition.pk)
        
        excel_file = request.FILES['excel_file']
        
        try:
            # Читаем Excel файл
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            created_count = 0
            updated_count = 0
            errors = []
            
            # Обрабатываем каждую строку, начиная с третьей
            for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
                if not row or row[0] is None:  # Пропускаем пустые строки
                    continue
                
                try:
                    sample_id = int(row[0])
                    
                    # Находим пробу
                    sample = Sample.objects.get(
                        sample_id=sample_id,
                        station__expedition=expedition
                    )
                    
                    # Создаем или обновляем метеоданные
                    meteo_data, created = MeteoData.objects.get_or_create(sample=sample)
                    
                    # Заполняем поля (колонки 5-9)
                    if row[4] is not None:  # Температура воздуха
                        meteo_data.t_air_c = float(row[4])
                    if row[5] is not None:  # Влажность
                        meteo_data.humidity_percent = float(row[5])
                    if row[6] is not None:  # Скорость ветра
                        meteo_data.wind_speed_m_s = float(row[6])
                    if row[7] is not None:  # Направление ветра
                        meteo_data.wind_direction = int(row[7])
                    if row[8] is not None:  # Давление
                        meteo_data.pressure_hpa = float(row[8])
                    
                    meteo_data.save()
                    
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                        
                except Sample.DoesNotExist:
                    errors.append(f"Строка {row_num}: Проба с ID {row[0]} не найдена в экспедиции")
                except ValueError as e:
                    errors.append(f"Строка {row_num}: Ошибка преобразования данных - {str(e)}")
                except Exception as e:
                    errors.append(f"Строка {row_num}: Неизвестная ошибка - {str(e)}")
            
            # Формируем сообщения о результате
            if created_count > 0 or updated_count > 0:
                success_msg = f"Успешно обработано: {created_count} новых записей, {updated_count} обновлений"
                messages.success(request, success_msg)
            
            if errors:
                error_msg = "Обнаружены ошибки: " + "; ".join(errors[:5])
                if len(errors) > 5:
                    error_msg += f" ... и еще {len(errors) - 5} ошибок"
                messages.error(request, error_msg)
            elif created_count == 0 and updated_count == 0:
                messages.warning(request, "Не было обработано ни одной записи. Проверьте формат файла.")
                
        except Exception as e:
            messages.error(request, f"Ошибка при обработке файла: {str(e)}")
        
        return redirect('oceanography:add_meteo_excel', expedition_id=expedition.pk)

        
class StationExcelUploadView(View):
    """Массовое добавление станций и проб через Excel"""
    template_name = 'oceanography/station_excel_upload.html'
    
    def get(self, request, *args, **kwargs):
        expedition_id = self.kwargs.get('expedition_id')
        expedition = get_object_or_404(Expedition, pk=expedition_id)
        
        context = {
            'expedition': expedition,
            'breadcrumbs': [
                {'url': reverse('oceanography:home'), 'name': 'Главная'},
                {'url': reverse('oceanography:expedition_list'), 'name': 'Экспедиции'},
                {'url': reverse('oceanography:expedition_detail', kwargs={'pk': expedition_id}), 
                 'name': f'Экспедиция {expedition.platform}'},
                {'url': '', 'name': 'Массовое добавление станций (Excel)'}
            ]
        }
        return render(request, self.template_name, context)
    
    def post(self, request, *args, **kwargs):
        expedition_id = self.kwargs.get('expedition_id')  # Получаем expedition_id из URL
        expedition = get_object_or_404(Expedition, pk=expedition_id)
        
        action = request.POST.get('action')
        
        if action == 'download_template':
            return self.download_template(expedition)
        elif action == 'upload_data':
            return self.upload_data(request, expedition)
        
        # Используем expedition_id из kwargs, а не неопределенную переменную
        return redirect('oceanography:add_stations_excel', expedition_id=self.kwargs.get('expedition_id'))
    
    def download_template(self, expedition):
        """Генерация шаблона Excel для массового добавления станций с пробами"""
        # Создаем новую книгу Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Станции и пробы"
        
        # Заголовки для станций (машинные и человекочитаемые)
        station_headers_machine = [
            'station_name', 'datetime', 'latitude', 'longitude', 
            'bottom_depth', 'secchi_depth'
        ]
        station_headers_human = [
            'Название станции*', 'Дата и время станции* (ГГГГ-ММ-ДД ЧЧ:ММ:СС)', 
            'Широта*', 'Долгота*', 'Глубина дна (м)', 'Прозрачность по Секки (м)'
        ]
        
        # Заголовки для проб (машинные и человекочитаемые)
        sample_headers_machine = ['sample_datetime', 'sampling_depth', 'sample_comment']
        sample_headers_human = ['Дата и время пробы', 'Горизонт отбора*', 'Комментарий к пробе']
        
        # Создаем заголовки: сначала данные станции, затем два набора проб
        headers_machine = station_headers_machine.copy()
        headers_human = station_headers_human.copy()
        
        for i in range(1, 3):  # 2 набора проб по умолчанию
            for header in sample_headers_machine:
                headers_machine.append(f'{header}_{i}')
            for header in sample_headers_human:
                suffix = f' (Проба {i})' if 'Дата' not in header else f' (Проба {i}, если отличается от станции)'
                headers_human.append(f'{header}{suffix}')
        
        # Записываем две строки заголовков
        for col, header in enumerate(headers_machine, 1):
            ws.cell(row=1, column=col, value=header)
        for col, header in enumerate(headers_human, 1):
            ws.cell(row=2, column=col, value=header)
        
        # Пример данных
        example_stations = [
            {
                'station_name': 'СТАНЦИЯ_1',
                'datetime': '2024-07-15 10:00:00',
                'latitude': 55.123456,
                'longitude': 37.654321,
                'bottom_depth': 25.5,
                'secchi_depth': 12.3,
                'samples': [
                    {
                        'datetime': '2024-07-15 10:00:00', 
                        'depth': '0.0', 
                        'comment': 'Поверхностная проба'
                    },
                    {
                        'datetime': '2024-07-15 10:30:00', 
                        'depth': 'дно', 
                        'comment': 'Придонная проба'
                    }
                ]
            },
            {
                'station_name': 'СТАНЦИЯ_2', 
                'datetime': '2024-07-15 12:00:00',
                'latitude': 55.223456,
                'longitude': 37.754321,
                'bottom_depth': 30.0,
                'secchi_depth': 10.0,
                'samples': [
                    {
                        'datetime': '2024-07-15 12:00:00', 
                        'depth': '0.0', 
                        'comment': 'Поверхностная проба'
                    },
                    {
                        'datetime': '2024-07-15 12:20:00', 
                        'depth': '15.0', 
                        'comment': 'Проба на 15 метрах'
                    }
                ]
            }
        ]
        
        # Заполняем примеры данных (начинаем с 3-й строки)
        for row, station_data in enumerate(example_stations, 3):
            # Данные станции
            ws.cell(row=row, column=1, value=station_data['station_name'])
            ws.cell(row=row, column=2, value=station_data['datetime'])
            ws.cell(row=row, column=3, value=station_data['latitude'])
            ws.cell(row=row, column=4, value=station_data['longitude'])
            ws.cell(row=row, column=5, value=station_data['bottom_depth'])
            ws.cell(row=row, column=6, value=station_data['secchi_depth'])
            
            # Данные проб
            sample_col = 7  # Начинаем с 7-й колонки
            for sample in station_data['samples']:
                ws.cell(row=row, column=sample_col, value=sample['datetime'])
                ws.cell(row=row, column=sample_col+1, value=sample['depth'])
                ws.cell(row=row, column=sample_col+2, value=sample['comment'])
                sample_col += 3
        
        # Стилизация заголовков
        for row in [1, 2]:
            for cell in ws[row]:
                cell.font = openpyxl.styles.Font(bold=True)
                if row == 1:
                    # Машинные заголовки - серый фон
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
                    )
                    cell.font = openpyxl.styles.Font(bold=True, color="666666")
                else:
                    # Человекочитаемые заголовки - синий фон
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color="E7F1FF", end_color="E7F1FF", fill_type="solid"
                    )
        
        # Добавляем пояснения на отдельный лист
        ws_help = wb.create_sheet("Инструкция")
        ws_help.append(["Поле", "Описание", "Обязательное", "Пример"])
        instructions = [
            ["station_name", "Уникальное название станции", "Да", "СТАНЦИЯ_1"],
            ["datetime", "Дата и время станции в формате ГГГГ-ММ-ДД ЧЧ:ММ:СС", "Да", "2024-07-15 10:00:00"],
            ["latitude", "Широта в десятичном формате", "Да", "55.123456"],
            ["longitude", "Долгота в десятичном формате", "Да", "37.654321"],
            ["bottom_depth", "Глубина дна в метрах", "Нет", "25.5"],
            ["secchi_depth", "Прозрачность по диску Секки в метрах", "Нет", "12.3"],
            ["sample_datetime_X", "Дата и время пробы X", "Нет", "2024-07-15 10:00:00"],
            ["sampling_depth_X", "Горизонт отбора пробы X", "Да", "0.0, 10.0, дно, поверхность"],
            ["sample_comment_X", "Комментарий к пробе X", "Нет", "Поверхностная проба"],
        ]
        
        for instruction in instructions:
            ws_help.append(instruction)
        
        # Автоподбор ширины колонок для основного листа
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)  # Ограничиваем максимальную ширину
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Автоподбор для листа инструкции
        for column in ws_help.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws_help.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Создаем HTTP response
        filename = f"station_template_{expedition.platform}_{expedition.start_date.year}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    def upload_data(self, request, expedition):
        """Обработка загруженного Excel файла со станциями и пробами"""
        if 'excel_file' not in request.FILES:
            messages.error(request, 'Пожалуйста, выберите файл для загрузки')
            return redirect('oceanography:add_stations_excel', expedition_id=expedition.pk)
        
        excel_file = request.FILES['excel_file']
        
        try:
            # Читаем Excel файл
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            created_stations = 0
            created_samples = 0
            errors = []
            
            # Обрабатываем каждую строку, начиная с ТРЕТЬЕЙ (первые две - заголовки)
            for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
                if not row or row[0] is None:  # Пропускаем пустые строки
                    continue
                
                try:
                    # Читаем данные станции
                    station_data = {
                        'station_name': row[0],
                        'datetime': row[1],
                        'latitude': row[2],
                        'longitude': row[3],
                        'bottom_depth': row[4],
                        'secchi_depth': row[5]
                    }
                    
                    # Проверяем обязательные поля
                    if not all([station_data['station_name'], station_data['datetime'], 
                               station_data['latitude'] is not None, station_data['longitude'] is not None]):
                        errors.append(f"Строка {row_num}: Отсутствуют обязательные данные станции")
                        continue
                    
                    # Проверяем уникальность станции
                    if Station.objects.filter(
                        expedition=expedition,
                        station_name=station_data['station_name'],
                        datetime=station_data['datetime']
                    ).exists():
                        errors.append(f"Строка {row_num}: Станция с таким названием и датой уже существует")
                        continue
                    
                    # Создаем станцию
                    station = Station.objects.create(
                        expedition=expedition,
                        **{k: v for k, v in station_data.items() if v is not None}
                    )
                    created_stations += 1
                    
                    # Читаем данные проб (колонки 6+)
                    sample_col = 6
                    sample_num = 1
                    
                    # Обрабатываем максимум 2 пробы (6 колонок)
                    while sample_col < len(row) and sample_col <= 11:  # 6 + 2*3 = 12, но индекс с 0
                        sample_data = {
                            'datetime': row[sample_col],
                            'sampling_depth': row[sample_col + 1],
                            'comment': row[sample_col + 2]
                        }
                        
                        # Если горизонт отбора указан, создаем пробу
                        if sample_data['sampling_depth']:
                            if not sample_data['datetime']:
                                sample_data['datetime'] = station.datetime
                            if not sample_data['comment']:
                                sample_data['comment'] = f'Проба {sample_num}'
                            
                            Sample.objects.create(
                                station=station,
                                **{k: v for k, v in sample_data.items() if v is not None}
                            )
                            created_samples += 1
                        
                        sample_col += 3
                        sample_num += 1
                        
                except Exception as e:
                    errors.append(f"Строка {row_num}: Ошибка обработки - {str(e)}")
            
            # Формируем сообщения о результате
            if created_stations > 0:
                success_msg = f"Успешно создано: {created_stations} станций, {created_samples} проб"
                messages.success(request, success_msg)
            
            if errors:
                error_msg = "Обнаружены ошибки: " + "; ".join(errors[:5])
                if len(errors) > 5:
                    error_msg += f" ... и еще {len(errors) - 5} ошибок"
                messages.error(request, error_msg)
            elif created_stations == 0:
                messages.warning(request, "Не было создано ни одной станции. Проверьте формат файла.")
                
        except Exception as e:
            messages.error(request, f"Ошибка при обработке файла: {str(e)}")
        
        return redirect('oceanography:add_stations_excel', expedition_id=expedition.pk)